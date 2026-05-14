import frappe
import csv
import openpyxl
from io import StringIO, BytesIO

def validate_bulk_import_payload(data):
    doctype = data.get("doctype")
    records = data.get("data")
    if not doctype:
        raise frappe.ValidationError("The 'doctype' field is required.")
    if not frappe.db.exists("DocType", doctype):
        raise frappe.ValidationError(f"DocType '{doctype}' does not exist.")
    if not records or not isinstance(records, list):
        raise frappe.ValidationError("The 'data' field must be a non-empty array of records.")

def get_doctype_schema(doctype):
    meta = frappe.get_meta(doctype)
    ignored_fieldtypes = ["Section Break", "Column Break", "HTML", "Button", "Fold", "Image", "Table"]
    schema = []
    schema.append({
        "fieldname": "name",
        "label": "ID (Leave blank to create new)",
        "reqd": 0,
        "fieldtype": "Data"
    })
    for df in meta.fields:
        if df.fieldtype not in ignored_fieldtypes and not df.read_only:
            schema.append({
                "fieldname": df.fieldname,
                "label": df.label,
                "reqd": df.reqd,
                "fieldtype": df.fieldtype,
                "options": df.options,
                "default": df.default
            })
    return schema

def validate_file_upload_payload(doctype, file_doc):
    if not doctype:
        raise frappe.ValidationError("Doctype is required.")
    if not file_doc or getattr(file_doc, "filename", None) is None:
        raise frappe.ValidationError("A valid file must be uploaded.")
    allowed_extensions = (".csv", ".xlsx", ".xls")
    if not file_doc.filename.endswith(allowed_extensions):
        raise frappe.ValidationError(f"Only {', '.join(allowed_extensions)} files are supported.")

def extract_records_from_file(filename, file_content):
    records = []
    if filename.endswith(".csv"):
        content_str = file_content.decode('utf-8-sig')
        reader = csv.DictReader(StringIO(content_str))
        records = [row for row in reader if any(row.values())]
    elif filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else f"Column_{i}" for i, cell in enumerate(ws[1])]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            if any(val is not None and str(val).strip() != "" for val in row_dict.values()):
                records.append(row_dict)
    if not records:
        raise frappe.ValidationError("The uploaded file contains no valid data rows.")
    return records